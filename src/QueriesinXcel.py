# from datetime import date

# Creating an empty workbook
# new_workbook = openpyxl.Workbook()
# load_workbook = openpyxl.load_workbook('/home/kirankumar/Shared/Common Files/Queries_Used.xlsx')
# date = date.today()
# Creating a workbook from a file
# workbook_from_file = load_workbook('/home/kirankumar/Shared/Common Files/Queries_Used.xlsx')
# Creating a sheet in the workbook
# newSheet = new_workbook.create_sheet('Queries')
# To copy the contents from one sheet to another sheet
# copy_sheet = new_workbook.copy_worksheet('sheet2')
# default_sheet = new_workbook.active
# default_sheet.title = "Queries"
# print (newSheet.title)
# newSheet.title = 'Queries'
# print (newSheet.title)
# sheet = new_workbook['Queries']
# for sheet in new_workbook:
#     print (sheet.title)
# print (default_sheet.title)
# c1 = default_sheet.cell(row=1, column=1)
# c1.value = 'Date'
# c2 = default_sheet.cell(row=1, column=2)
# c2.value = 'Query'
# c3 = default_sheet.cell(row=1, column=3)
# c3.value = 'Description'
# new_workbook.save('/home/kirankumar/Shared/Common Files/Queries_Used.xlsx')
# print ("Entries are added and excel is saved")

# filePath = '/home/kirankumar/Shared/Common Files/Queries_Used.xlsx'
# workbook =
# sheet = load.active
# print (sheet.title)
# cell_range = sheet['A':'C']
# print ("The loaded sheet contains {} rows and {} columns".format(sheet.max_row, sheet.max_column))
# query = raw_input("Please enter the query: ")
# desc = raw_input("Please enter the description: ")
# data = [('date', 'query', 'desc')]
# for row in data:
#     sheet.append(row)
# load.save(load)

import openpyxl
from datetime import date

filePath = '/home/kirankumar/Shared/Common Files/Queries_Used.xlsx'
workbook = openpyxl.load_workbook(filePath)
sheet = workbook.active
print ("Working sheet is: {}".format(sheet.title))
date = date.today()
query = raw_input("Please enter the Query: ")
desc = raw_input("Please enter the description of that query: ")
data = [(date, query, desc)]

for row in data:
    sheet.append(row)

workbook.save(filePath)
